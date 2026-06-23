#!/usr/bin/env python3
"""
edit_menu_excel.py
==================
Post-process a finished menu workbook (the .xlsx produced by the scanner) and
apply structural edits in place:

Menu sheet
  * add "Category Optional Modifier ID" between "Category 3" and "Item ID"
  * add "Item Optional Modifier ID", "Picture", "Description" between "Price"
    and "Sec"

Section sheet
  * write each section name once: keep it on the top row of a run and blank the
    repeated name cells below it (the name applies downward until a different
    name appears). No rows are deleted.

New "Optional Modifiers" sheet
  * created with columns: Modified Category ID, Modified Category, Name,
    Modifier ID, Modifier Name, Price Open Price

The script is idempotent — running it twice will not add the columns/sheet again.

Usage
-----
    python edit_menu_excel.py --input menu.xlsx               # edit in place
    python edit_menu_excel.py --input menu.xlsx --output edited.xlsx
"""

import argparse
import sys

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NEW_OPTMOD_SHEET = "Optional Modifiers"
OPTMOD_COLUMNS = ["Modified Category ID", "Modified Category", "Name",
                  "Modifier ID", "Modifier Name", "Price Open Price"]

_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", start_color="2F5496")
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def style_header_cell(cell, value):
    cell.value = value
    cell.font = _HEADER_FONT
    cell.fill = _HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _BORDER


def col_index(ws, header_name):
    """1-based column index of a header on row 1, or None."""
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=c).value == header_name:
            return c
    return None


# --------------------------------------------------------------------------- #
# Menu sheet: insert the new columns
# --------------------------------------------------------------------------- #

def edit_menu_sheet(ws):
    added = []

    # Three columns between "Price" and "Sec". Do this first (it is to the right
    # of "Category 3", so the Category 3 index stays valid for the next step).
    if col_index(ws, "Item Optional Modifier ID") is None:
        price = col_index(ws, "Price")
        if price is None:
            print("  ! 'Price' column not found; skipping Item/Picture/Description.")
        else:
            ws.insert_cols(price + 1, amount=3)
            for offset, name in enumerate(
                    ["Item Optional Modifier ID", "Picture", "Description"], start=1):
                style_header_cell(ws.cell(row=1, column=price + offset), name)
                ws.column_dimensions[get_column_letter(price + offset)].width = 22
            added += ["Item Optional Modifier ID", "Picture", "Description"]

    # One column between "Category 3" and "Item ID".
    if col_index(ws, "Category Optional Modifier ID") is None:
        cat3 = col_index(ws, "Category 3")
        if cat3 is None:
            print("  ! 'Category 3' column not found; skipping Category Optional Modifier ID.")
        else:
            ws.insert_cols(cat3 + 1, amount=1)
            style_header_cell(ws.cell(row=1, column=cat3 + 1), "Category Optional Modifier ID")
            ws.column_dimensions[get_column_letter(cat3 + 1)].width = 24
            added.append("Category Optional Modifier ID")

    if added:
        print(f"  Menu: added columns {added}")
    else:
        print("  Menu: columns already present, nothing to add.")


# --------------------------------------------------------------------------- #
# Section sheet: keep only the topmost group for each SectionName
# --------------------------------------------------------------------------- #

def collapse_section_names(ws):
    """Write each section name once: on the first row of a run, leave the cells
    below it blank. The name applies downward until a different name appears
    (same convention as the Category column on the Menu sheet). No rows removed.
    """
    name_idx = col_index(ws, "SectionName")
    if name_idx is None:
        print("  ! Section sheet missing 'SectionName'; skipping collapse.")
        return

    current = None
    blanked = 0
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=name_idx).value
        if name is None:
            continue                      # already blank -> part of current section
        if name == current:
            ws.cell(row=r, column=name_idx).value = None   # repeat -> blank it
            blanked += 1
        else:
            current = name                # new section name -> keep it written

    if blanked:
        print(f"  Section: collapsed repeated section names ({blanked} cell(s) blanked).")
    else:
        print("  Section: section names already written once each.")


# --------------------------------------------------------------------------- #
# Add the "Optional Modifiers" sheet
# --------------------------------------------------------------------------- #

def add_optional_modifiers_sheet(wb):
    if NEW_OPTMOD_SHEET in wb.sheetnames:
        print(f"  '{NEW_OPTMOD_SHEET}' sheet already exists, leaving it as is.")
        return
    ws = wb.create_sheet(NEW_OPTMOD_SHEET)
    for c, name in enumerate(OPTMOD_COLUMNS, start=1):
        style_header_cell(ws.cell(row=1, column=c), name)
        ws.column_dimensions[get_column_letter(c)].width = max(14, len(name) + 2)
    ws.freeze_panes = "A2"
    print(f"  Added '{NEW_OPTMOD_SHEET}' sheet with columns {OPTMOD_COLUMNS}")


# --------------------------------------------------------------------------- #

def main():
    ap = argparse.ArgumentParser(description="Apply structural edits to a finished menu workbook.")
    ap.add_argument("--input", required=True, help="Path to the .xlsx to edit.")
    ap.add_argument("--output", help="Where to write (default: overwrite --input in place).")
    args = ap.parse_args()

    out = args.output or args.input
    try:
        wb = load_workbook(args.input)
    except Exception as exc:
        sys.exit(f"Could not open {args.input}: {exc}")

    if "Menu" in wb.sheetnames:
        edit_menu_sheet(wb["Menu"])
    else:
        print("  ! No 'Menu' sheet found.")

    if "Section" in wb.sheetnames:
        collapse_section_names(wb["Section"])
    else:
        print("  ! No 'Section' sheet found.")

    add_optional_modifiers_sheet(wb)

    wb.save(out)
    print(f"Saved -> {out}")


if __name__ == "__main__":
    main()
