#!/usr/bin/env python3
"""
menu_to_excel.py
================
Scan a menu image, analyze it, and write a structured Excel file.

Columns (matches the reference layout, with a single combined Sec column):
  Menu ID | Menu Group | Menu Group 2 | Menu Group 3 |
  Category ID | Category | Category 2 | Category 3 |
  Item ID | Item | Item 2 | Item 3 | Item Type | Price | Sec

Field rules
-----------
CATEGORY  Real printed sections only (Appetizers, Mains, Desserts, ...). The
          program never invents a category and never moves an item out of its
          printed section. Category 2 / Category 3 are filled ONLY when the
          section heading is actually printed in a second / third language.

ITEM      Item 2 / Item 3 hold the item name in another language, and only when
          that translation is actually printed.

ITEM TYPE 1 = FIXED        -> always costs the listed price.
          2 = MODIFIABLE    -> price varies; use only when obvious, e.g. an
                               "S.P." / "M.P." / market-price marker.
          3 = MULTI-AS-ONE  -> may order several but they count as one line item
                               (e.g. 5 donuts still count as one item).

SEC       A combinability group. A sec id marks a set of items that can be
          combined into one meal/combo. If a pork-over-rice can be combined with
          a soup and a drink, those items share that combo's id. An item that
          can go into several different combos lists several ids. A pre-set
          bundle (e.g. "Dinner for 4") is a single item row with its own unique
          id - its contents are NOT listed as separate items. The Sec column is
          a single comma-separated list, e.g. "1,2,3". Items not part of any
          combo have an empty Sec. Combos are expressed ONLY here, never as
          their own category.

Engine
------
Plain OCR cannot judge combinability or item type, so the analysis is done by
Claude's vision model: the image is sent to the Anthropic Messages API, which
returns strict JSON that is then mapped onto the spreadsheet.

Usage
-----
    export ANTHROPIC_API_KEY="sk-ant-..."
    python menu_to_excel.py --image menu.jpg --output menu.xlsx
    python menu_to_excel.py --folder ./menu_pages --output menu.xlsx   # whole menu
    python menu_to_excel.py --image menu.jpg --output menu.xlsx --save-json out.json
    python menu_to_excel.py --from-json out.json --output menu.xlsx    # offline rebuild

A folder is treated as ONE menu: every image in it is sent together (in natural
page order, so page2 before page10), so sections that span pages stay merged and
Item IDs run continuously.

Requirements
------------
    pip install anthropic openpyxl
"""

import argparse
import base64
import json
import mimetypes
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

MODEL = "claude-opus-4-8"
MAX_TOKENS = 16000

# Image extensions recognized when scanning a folder of menu pages.
IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".gif"}

# USD per 1,000,000 tokens, as (input, output). Edit if rates change.
PRICES = {
    "claude-opus-4-8":            (5.0, 25.0),
    "claude-sonnet-4-6":          (3.0, 15.0),
    "claude-haiku-4-5-20251001":  (1.0,  5.0),
}
DEFAULT_PRICE = (5.0, 25.0)

# Running token total for this process (one API call per run here).
USAGE = {"input_tokens": 0, "output_tokens": 0}


def estimate_cost(model: str, input_tokens: int, output_tokens: int) -> float:
    price_in, price_out = PRICES.get(model, DEFAULT_PRICE)
    return input_tokens / 1_000_000 * price_in + output_tokens / 1_000_000 * price_out

COLUMNS = [
    "Menu ID", "Menu Group", "Menu Group 2", "Menu Group 3",
    "Category ID", "Category", "Category 2", "Category 3",
    "Item ID", "Item", "Item 2", "Item 3", "Item Type", "Price", "Sec",
]

# Second sheet: one row per size variant, linked to the Menu sheet by Item id.
VARIANT_COLUMNS = ["Item id", "variant id", "variant name", "Language 2 name", "price"]

# Third sheet: one row per modifier option (choices / add-ons) that apply to a
# whole section/category, e.g. the rice-or-noodle choice on a Hibachi Meal.
SECTION_COLUMNS = ["Section Id", "SectionName", "Min", "Max",
                   "Modifier ID", "ModifierName", "ModifierName2", "Price"]


# --------------------------------------------------------------------------- #
# Step 1 — extraction prompt (the heart of the analysis)
# --------------------------------------------------------------------------- #

def build_prompt(n_pages: int = 1) -> str:
    if n_pages > 1:
        preamble = (
            f"You are reading ONE restaurant menu that has been photographed as "
            f"{n_pages} images, given in page order. Treat them as a single menu: "
            f"combine everything into one result, do not restart anything per page, "
            f"and if a section continues onto the next page do NOT repeat its "
            f"heading as a new section.\n"
        )
    else:
        preamble = "You are reading a restaurant menu from the attached image.\n"
    return preamble + """Extract every menu item and return ONE JSON object and nothing else
(no prose, no markdown fences).

JSON shape:
{
  "menu_name": "name of the menu or restaurant if visible, else null",
  "menu_name_translations": [],
  "combos": [
    {"id": 1, "label": "short human name of a combo / meal deal printed on the menu"}
  ],
  "sections": [
    {
      "name": "section heading exactly as printed (e.g. Appetizers, Soups)",
      "name_translations": [],
      "modifier_groups": [],
      "items": [
        {
          "name_primary": "item name in the main language, exactly as printed",
          "name_translations": [],
          "price": 8.95,
          "item_type": 1,
          "sec": [],
          "variants": []
        }
      ]
    }
  ]
}

Rules:

1. SECTIONS / CATEGORY: Use only sections that are actually printed on the menu
   (Appetizers, Soups, Mains, Desserts, etc.). Do NOT invent sections. Never
   create a "Combo" or "Combination" section just because items can be combined
   - keep each item in the section it is printed under. Keep the printed order.

2. TRANSLATIONS: "name_translations" (for sections and items) and
   "menu_name_translations" must ONLY contain a name in another language when
   that translation is actually printed on the menu. If only one language is
   printed, leave these lists empty. List them in printed order (second
   language first, then third).

3. PRICE: a number only (e.g. 8.95). If there is no fixed numeric price
   (e.g. "S.P." / "M.P." / market price, or a combo header with no standalone
   price) use 0 and set item_type accordingly.

4. ITEM TYPE — exactly one integer:
     1 = FIXED: normal item, always costs its listed price.
     2 = MODIFIABLE PRICE: use ONLY when obvious, e.g. the item is marked
         "S.P." / "M.P." / "market price" / "seasonal price". When unsure use 1.
     3 = MULTI-AS-ONE: you can order several but they still count as one line
         item (e.g. donuts sold per piece where 5 donuts count as 1 item).

5. SEC (combinability groups) — this is the important one. Each distinct combo
   gets one numeric id in the top-level "combos" list, and "sec" on an item is
   the list of combo ids that item belongs to. There are two kinds of combo:

   (a) MIX-AND-MATCH combo: the menu lets you assemble a meal by choosing items
       ("choice of", "served with", a lunch special with a list of entrees).
       The selectable items are real a-la-carte items that appear in their own
       sections; tag each of them with that combo's SHARED id. An item usable
       in several such combos lists several ids, e.g. [1, 2].
       Example: a "Pork Over Rice" lunch box (combo id 1) comes with a choice of
       soup and a drink. Then Pork Over Rice -> [1]; each eligible soup -> id 1;
       each eligible drink -> id 1. If a soup is also offered in a dinner combo
       (id 2), that soup -> [1, 2].

   (b) PRE-SET BUNDLE / SET MEAL: a single fixed thing sold as one unit whose
       contents are not individually chosen (e.g. "Dinner for 4", "Family
       Special", a fixed appetizer platter). Output it as ONE item row using the
       bundle's name. Do NOT create separate item rows for the things inside it
       (you may keep the contents in parentheses in the name if printed, but do
       not list them as items). Add the bundle to "combos" with its own id and
       set that one item's sec to that id alone — a UNIQUE id no other item uses.

   - Do NOT give ordinary items their own id. ids exist only for real combos.
   - Items not part of any combo have sec [].

6. SIZE VARIANTS vs DIFFERENT DISHES:
   A "variant" is the SAME identical dish offered in different SIZES / PORTIONS
   only - same food, just a bigger or smaller portion (e.g. "Pt." / "Qt.",
   "Small" / "Large", "Sm" / "Lg"). List those in "variants":
     "variants": [
       {"name": "Pt.", "name_translation": null, "price": 6.50},
       {"name": "Qt.", "name_translation": null, "price": 9.95}
     ]
   - "name" is the size label exactly as printed (e.g. "Pt.", "Qt.", "S", "L").
   - "name_translation" is that size label in the menu's other language if one is
     actually printed, else null.
   - "price" is that size's price.
   - Set the item's top-level "price" to the FIRST (usually smallest) variant's
     price, and still list EVERY size (including the first) in "variants".

   If the differently-priced options change WHAT the dish IS - a different
   accompaniment, side, protein, or ingredient (e.g. "w. Plain" vs "w. Roast Pork
   Fried Rice" vs "w. Shrimp or Beef Fried Rice", or "w. Chicken" vs "w. Beef") -
   they are DIFFERENT DISHES, not variants. Make EACH option its own separate
   item with its own price, naming it by combining the base dish with the option
   (e.g. "Fried Chicken Wings (4) w. Roast Pork Fried Rice"). Give each such item
   "variants": []. Do NOT put these options in "variants".

   - Items sold at a single price have "variants": [] (do not invent sizes).

7. SPLIT "EITHER/OR" ITEMS: When one menu line offers interchangeable
   alternatives joined by "or", output a SEPARATE item for EACH alternative,
   distributing the shared words (which may come before or after the "or"):
     "Egg Roll or Spring Roll"        -> "Egg Roll" + "Spring Roll"
     "Fried or Steamed Dumpling"      -> "Fried Dumpling" + "Steamed Dumpling"
     "Roast Pork or Chicken Lo Mein"  -> "Roast Pork Lo Mein" + "Chicken Lo Mein"
     "Sweet & Sour Chicken or Pork"   -> "Sweet & Sour Chicken" + "Sweet & Sour Pork"
     "Roast Pork or Chicken or Vegetable Fried Rice" -> three separate items
   Each split item is its own item and keeps the SAME price, item_type, sec, and
   variants as the original line, plus the same leading menu number/letter if any
   (e.g. "L6. Sweet & Sour Chicken" and "L6. Sweet & Sour Pork"). Split the
   second-language name the same way if it also lists the alternatives (e.g. with
   "或"); otherwise repeat the same translation on each. Only split genuine
   either/or choices — do not split an "or" that is part of one dish's description.

8. SECTION MODIFIERS (choices / add-ons that apply to a whole section): A section
   heading sometimes says how every item in it is ordered — a required choice or
   optional add-ons. Capture these per section in "modifier_groups". Examples:
     "Served w. choice of Steamed Rice, Fried Rice (Noodles Extra $1.00, brown
      rice Extra $1.00)"  -> a REQUIRED choice group (min 1, max 1) with options
      Steamed Rice (price 0), Fried Rice (price 0), Noodles (price 1.00),
      Brown Rice (price 1.00).
     "Add to your Hibachi: *Chicken 2.99 *Beef 2.99 *Shrimp 2.99 *Salmon 4.00"
      -> an OPTIONAL add-on group (min 0, max 1) with each protein and its price.
   Shape per section:
     "modifier_groups": [
       {"min": 1, "max": 1, "options": [
          {"name": "Steamed Rice", "name_translation": null, "price": 0},
          {"name": "Noodles", "name_translation": null, "price": 1.00}]}
     ]
   - "price" is the SURCHARGE for that option (0 if included at no extra cost,
     otherwise the extra amount such as 1.00 or 2.99).
   - "min"/"max" are how many options must / may be chosen from that group. A
     required "choice of" is min 1, max 1. Optional add-ons are min 0, max 1.
     Cap both at 1 unless the menu clearly allows more.
   - name_translation is the option name in the other language if printed, else null.
   - Sections with no such choices have "modifier_groups": [].

Return only the JSON object."""


# --------------------------------------------------------------------------- #
# Step 2 — call the vision model
# --------------------------------------------------------------------------- #

def encode_image(path: str):
    media_type, _ = mimetypes.guess_type(path)
    if media_type is None or not media_type.startswith("image/"):
        media_type = "image/jpeg"
    with open(path, "rb") as fh:
        data = base64.standard_b64encode(fh.read()).decode("utf-8")
    return data, media_type


def natural_key(path: str):
    """Sort key so page2 comes before page10 (handles numbers in filenames)."""
    import re
    name = os.path.basename(path).lower()
    return [int(t) if t.isdigit() else t for t in re.split(r"(\d+)", name)]


def list_images(folder: str) -> list:
    """Return the image files in `folder`, sorted in natural page order."""
    if not os.path.isdir(folder):
        sys.exit(f"Not a folder: {folder}")
    paths = [
        os.path.join(folder, name)
        for name in os.listdir(folder)
        if os.path.splitext(name)[1].lower() in IMAGE_EXTS
    ]
    if not paths:
        sys.exit(f"No image files ({', '.join(sorted(IMAGE_EXTS))}) found in {folder}")
    return sorted(paths, key=natural_key)


def extract_with_claude(image_paths: list) -> dict:
    """Send one or more page images (in order) as a single menu to Claude."""
    try:
        from anthropic import Anthropic
    except ImportError:
        sys.exit("The 'anthropic' package is required. Run: pip install anthropic")
    if not os.environ.get("ANTHROPIC_API_KEY"):
        sys.exit("Set ANTHROPIC_API_KEY in your environment before running.")

    content = []
    for i, path in enumerate(image_paths, start=1):
        data, media_type = encode_image(path)
        if len(image_paths) > 1:
            content.append({"type": "text", "text": f"--- Page {i} ---"})
        content.append({
            "type": "image",
            "source": {"type": "base64", "media_type": media_type, "data": data},
        })
    content.append({"type": "text", "text": build_prompt(len(image_paths))})
    user_msg = {"role": "user", "content": content}

    continue_instruction = (
        "Your previous reply was cut off because it was too long. Continue the "
        "JSON from EXACTLY where it stopped — your next characters must directly "
        "follow the last characters of your previous reply. Do not repeat anything "
        "you already wrote, do not restart, and do not add any explanation or "
        "markdown fences. Output only the remaining JSON."
    )

    client = Anthropic()
    accumulated = ""
    for round_no in range(12):                 # safety cap on continuations
        if not accumulated:
            messages = [user_msg]
        else:
            # keep the partial reply as context, then end with a USER message
            # (this model does not allow ending on an assistant message).
            messages = [
                user_msg,
                {"role": "assistant", "content": accumulated},
                {"role": "user", "content": continue_instruction},
            ]
        message = client.messages.create(
            model=MODEL,
            max_tokens=MAX_TOKENS,
            messages=messages,
        )
        usage = getattr(message, "usage", None)
        if usage is not None:
            USAGE["input_tokens"] += getattr(usage, "input_tokens", 0) or 0
            USAGE["output_tokens"] += getattr(usage, "output_tokens", 0) or 0
        chunk = "".join(b.text for b in message.content if b.type == "text")
        # JSON ignores inter-token whitespace; trailing whitespace would also make
        # an invalid assistant turn, so trim it before the next round.
        accumulated = _stitch_text(accumulated, chunk).rstrip()
        if message.stop_reason != "max_tokens":
            break
        if round_no >= 1:
            print(f"  (page is long — continuing extraction, part {round_no + 2})")
    else:
        sys.exit("Page was too long to finish even after many continuation rounds.")

    return parse_json(accumulated)


def _stitch_text(accumulated: str, chunk: str) -> str:
    """Append a continuation chunk to the accumulated text, removing any overlap
    where the model re-emitted the boundary, and stray leading code fences."""
    chunk = chunk.lstrip()
    if chunk.startswith("```"):
        chunk = chunk.split("```", 2)[-1].lstrip()
        if chunk.lower().startswith("json"):
            chunk = chunk[4:]
    if not accumulated:
        return chunk
    # drop the largest suffix of `accumulated` that the chunk repeats as its prefix
    max_k = min(len(accumulated), len(chunk), 2000)
    for k in range(max_k, 9, -1):
        if accumulated[-k:] == chunk[:k]:
            chunk = chunk[k:]
            break
    return accumulated + chunk


def parse_json(text: str) -> dict:
    cleaned = text.strip()
    if cleaned.startswith("```"):
        cleaned = cleaned.split("```", 2)[1]
        if cleaned.lstrip().lower().startswith("json"):
            cleaned = cleaned.lstrip()[4:]
    cleaned = cleaned.strip().strip("`").strip()
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError as exc:
        sys.exit(f"Could not parse model output as JSON: {exc}\n---\n{text[:2000]}")


# --------------------------------------------------------------------------- #
# Step 3 — map JSON onto the spreadsheet schema
# --------------------------------------------------------------------------- #

def _as_list(value):
    """Normalize a translations field that may be None, a string, or a list."""
    if value is None:
        return []
    if isinstance(value, str):
        return [value] if value.strip() else []
    return [v for v in value if v]


def _sec_string(sec):
    """Turn a list of combo ids into '1,2,3' (or None when empty)."""
    if not sec:
        return None
    if isinstance(sec, (int, str)):
        sec = [sec]
    ids = []
    for s in sec:
        try:
            ids.append(int(s))
        except (TypeError, ValueError):
            continue
    return ",".join(str(i) for i in ids) if ids else None


def build_rows(menu: dict):
    """Return (menu_rows, variant_rows, section_rows).
    variant_rows link to menu rows by Item id; section_rows hold per-section
    modifier options (Section Id == the section's Category ID)."""
    rows = []
    variant_rows = []
    section_rows = []
    menu_name = menu.get("menu_name") or "Menu"
    menu_tr = _as_list(menu.get("menu_name_translations"))
    category_id = 0
    item_id = 0
    variant_id = 0
    modifier_id = 0
    first_row = True

    for section in menu.get("sections", []):
        category_id += 1
        sec_tr = _as_list(section.get("name_translations"))
        first_item = True
        for item in section.get("items", []):
            item_id += 1
            item_tr = _as_list(item.get("name_translations"))
            rows.append({
                "Menu ID":      1 if first_row else None,
                "Menu Group":   menu_name if first_row else None,
                "Menu Group 2": (menu_tr[0] if len(menu_tr) > 0 else None) if first_row else None,
                "Menu Group 3": (menu_tr[1] if len(menu_tr) > 1 else None) if first_row else None,
                "Category ID":  category_id if first_item else None,
                "Category":     section.get("name") if first_item else None,
                "Category 2":   (sec_tr[0] if len(sec_tr) > 0 else None) if first_item else None,
                "Category 3":   (sec_tr[1] if len(sec_tr) > 1 else None) if first_item else None,
                "Item ID":      item_id,
                "Item":         item.get("name_primary"),
                "Item 2":       item_tr[0] if len(item_tr) > 0 else None,
                "Item 3":       item_tr[1] if len(item_tr) > 1 else None,
                "Item Type":    int(item.get("item_type", 1) or 1),
                "Price":        round(float(item.get("price", 0) or 0), 2),
                "Sec":          _sec_string(item.get("sec")),
            })
            for variant in (item.get("variants") or []):
                variant_id += 1
                variant_rows.append({
                    "Item id":         item_id,
                    "variant id":      variant_id,
                    "variant name":    variant.get("name"),
                    "Language 2 name": variant.get("name_translation"),
                    "price":           round(float(variant.get("price", 0) or 0), 2),
                })
            first_row = False
            first_item = False

        # section-level modifier groups (rice/noodle choice, protein add-ons, ...)
        for group in (section.get("modifier_groups") or []):
            gmin = int(group.get("min", 0) or 0)
            gmax = int(group.get("max", 1) or 1)
            for opt in (group.get("options") or []):
                modifier_id += 1
                section_rows.append({
                    "Section Id":    category_id,
                    "SectionName":   section.get("name"),
                    "Min":           gmin,
                    "Max":           gmax,
                    "Modifier ID":   modifier_id,
                    "ModifierName":  opt.get("name"),
                    "ModifierName2": opt.get("name_translation"),
                    "Price":         round(float(opt.get("price", 0) or 0), 2),
                })

    return rows, variant_rows, section_rows


# --------------------------------------------------------------------------- #
# Step 4 — write the Excel file
# --------------------------------------------------------------------------- #

def write_excel(rows: list, output_path: str, variant_rows: list = None,
                section_rows: list = None) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Menu"

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="2F5496")
    base_font = Font(name="Arial")
    section_fill = PatternFill("solid", start_color="DCE6F1")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(sheet, columns):
        sheet.append(columns)
        for c in range(1, len(columns) + 1):
            cell = sheet.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        sheet.freeze_panes = "A2"

    # ---- Menu sheet ----
    style_header(ws, COLUMNS)
    center_cols = {"Menu ID", "Category ID", "Item ID", "Item Type", "Sec"}
    for r in rows:
        ws.append([r[c] for c in COLUMNS])
        excel_row = ws.max_row
        is_section_start = r["Category ID"] is not None
        for c, name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=excel_row, column=c)
            cell.font = base_font
            cell.border = border
            if name == "Price":
                cell.number_format = "0.00"
            if name == "Sec":
                cell.number_format = "@"   # keep "1,2,3" as text, never a number/date
            if name in center_cols:
                cell.alignment = Alignment(horizontal="center")
            if is_section_start:
                cell.fill = section_fill

    widths = {
        "Item": 42, "Item 2": 22, "Item 3": 18,
        "Category": 26, "Category 2": 20, "Category 3": 16,
        "Menu Group": 16, "Menu Group 2": 16, "Menu Group 3": 14, "Sec": 10,
    }
    for c, name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(c)].width = widths.get(name, 11)

    # ---- Variants sheet ----
    wsv = wb.create_sheet("Variants")
    style_header(wsv, VARIANT_COLUMNS)
    v_center = {"Item id", "variant id"}
    for r in (variant_rows or []):
        wsv.append([r.get(c) for c in VARIANT_COLUMNS])
        excel_row = wsv.max_row
        for c, name in enumerate(VARIANT_COLUMNS, start=1):
            cell = wsv.cell(row=excel_row, column=c)
            cell.font = base_font
            cell.border = border
            if name == "price":
                cell.number_format = "0.00"
            if name in v_center:
                cell.alignment = Alignment(horizontal="center")
    v_widths = {"Item id": 10, "variant id": 11, "variant name": 18, "Language 2 name": 18, "price": 11}
    for c, name in enumerate(VARIANT_COLUMNS, start=1):
        wsv.column_dimensions[get_column_letter(c)].width = v_widths.get(name, 12)

    # ---- Section sheet (per-section modifier options) ----
    wss = wb.create_sheet("Section")
    style_header(wss, SECTION_COLUMNS)
    s_center = {"Section Id", "Min", "Max", "Modifier ID"}
    for r in (section_rows or []):
        wss.append([r.get(c) for c in SECTION_COLUMNS])
        excel_row = wss.max_row
        for c, name in enumerate(SECTION_COLUMNS, start=1):
            cell = wss.cell(row=excel_row, column=c)
            cell.font = base_font
            cell.border = border
            if name == "Price":
                cell.number_format = "0.00"
            if name in s_center:
                cell.alignment = Alignment(horizontal="center")
    s_widths = {"Section Id": 10, "SectionName": 26, "Min": 7, "Max": 7,
                "Modifier ID": 11, "ModifierName": 20, "ModifierName2": 18, "Price": 11}
    for c, name in enumerate(SECTION_COLUMNS, start=1):
        wss.column_dimensions[get_column_letter(c)].width = s_widths.get(name, 12)

    wb.save(output_path)


# --------------------------------------------------------------------------- #
# Entry point
# --------------------------------------------------------------------------- #

def main() -> None:
    ap = argparse.ArgumentParser(description="Scan a menu (one image or a folder of pages) into a structured Excel file.")
    src = ap.add_mutually_exclusive_group(required=True)
    src.add_argument("--image", help="Path to a single menu image (jpg/png/webp/gif).")
    src.add_argument("--folder", help="Folder of menu page images, treated as ONE menu.")
    src.add_argument("--from-json", help="Previously extracted JSON (skips the API).")
    ap.add_argument("--output", required=True, help="Path for the .xlsx file to write.")
    ap.add_argument("--save-json", help="Optionally save the raw extracted JSON here.")
    args = ap.parse_args()

    output_path = args.output
    if not output_path.lower().endswith(".xlsx"):
        output_path = os.path.splitext(output_path)[0] + ".xlsx"
        print(f"Note: output is an Excel file, so writing it as {output_path} "
              f"(not {args.output}).")

    if args.from_json:
        with open(args.from_json, "r", encoding="utf-8") as fh:
            menu = json.load(fh)
    elif args.folder:
        images = list_images(args.folder)
        print(f"Scanning {len(images)} page(s) as one menu:")
        for p in images:
            print(f"  - {os.path.basename(p)}")
        menu = extract_with_claude(images)
    else:
        menu = extract_with_claude([args.image])

    if args.save_json:
        with open(args.save_json, "w", encoding="utf-8") as fh:
            json.dump(menu, fh, ensure_ascii=False, indent=2)

    rows, variant_rows, section_rows = build_rows(menu)
    write_excel(rows, output_path, variant_rows, section_rows)
    print(f"Wrote {len(rows)} items ({len(variant_rows)} size variants, "
          f"{len(section_rows)} section modifiers) to {output_path}")

    if not args.from_json:
        cost = estimate_cost(MODEL, USAGE["input_tokens"], USAGE["output_tokens"])
        pin, pout = PRICES.get(MODEL, DEFAULT_PRICE)
        print(f"Tokens: {USAGE['input_tokens']:,} in / {USAGE['output_tokens']:,} out")
        print(f"Estimated cost: ${cost:.4f}  (at ${pin}/{pout} per million in/out, {MODEL})")
        # machine-readable line for tools that run this script per page:
        print(f"__USAGE__ {json.dumps({'model': MODEL, **USAGE})}")


if __name__ == "__main__":
    main()
