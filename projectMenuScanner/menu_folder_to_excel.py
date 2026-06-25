#!/usr/bin/env python3
"""
menu_folder_to_excel.py
=======================
Turn a FOLDER of menu page photos into ONE Excel file.

Why this exists
---------------
Sending a whole multi-page menu in a single model call can overrun the model's
response limit, which truncates the JSON and breaks the run. This script instead
runs the existing per-image extractor (menu_to_excel_2.py) ONCE PER PAGE - each
page is small, so nothing gets cut off - and then stitches the resulting .xlsx
files together. The numeric merge is plain Python (openpyxl); the only AI in the
stitch step is one small, separate call that decides where a section was split
across a page break.

What the stitcher fixes up when concatenating pages
---------------------------------------------------
* Item ID      -> renumbered 1..N across the whole menu.
* Category ID  -> renumbered across the whole menu. A small, separate AI text
                  call decides, per page boundary, whether the section that
                  begins the next page is the SAME section still running at the
                  end of the previous page (a section split across the break);
                  if so it folds into one Category, even when the heading isn't
                  reprinted. (--no-detect falls back to matching section names.)
                  Pages are still read separately; this call sees only section
                  names and a few item names at each boundary, not the images.
* Sec          -> each page numbers its combos from 1, so every page's ids are
                  offset by the running total before merging (page 1 uses 1..k,
                  page 2's ids become k+1.., etc.) so combos never collide.
* Menu ID / Menu Group -> kept only on the very first row of the merged sheet.
All other columns (Item, translations, Item Type, Price) are carried across
unchanged.

Usage
-----
    export ANTHROPIC_API_KEY="sk-ant-..."
    python menu_folder_to_excel.py --folder ./menu_pages --output menu.xlsx

Options
-------
    --engine PATH     Path to the per-image extractor (default: menu_to_excel_2.py
                      next to this script).
    --keep-pages      Keep the intermediate per-page .xlsx files.
    --pages-dir DIR   Where to write the per-page files (default: a temp folder).
    --python EXE      Python interpreter used to run the engine (default: this one).

Requirements
------------
    pip install anthropic openpyxl   (same as the engine)
"""

import argparse
import importlib.util
import os
import subprocess
import sys
import tempfile

from openpyxl import load_workbook


# --------------------------------------------------------------------------- #
# Load the engine module (for its COLUMNS and write_excel) without running it
# --------------------------------------------------------------------------- #

def load_engine(engine_path: str):
    if not os.path.isfile(engine_path):
        sys.exit(f"Engine not found: {engine_path}")
    spec = importlib.util.spec_from_file_location("menu_engine", engine_path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)            # safe: engine guards its CLI with __main__
    return module


# --------------------------------------------------------------------------- #
# Find page images (natural order: page2 before page10)
# --------------------------------------------------------------------------- #

IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".gif"}


def natural_key(path: str):
    import re
    name = os.path.basename(path).lower()
    return [int(t) if t.isdigit() else t for t in re.split(r"(\d+)", name)]


def list_images(folder: str) -> list:
    if not os.path.isdir(folder):
        sys.exit(f"Not a folder: {folder}")
    paths = [
        os.path.join(folder, n)
        for n in os.listdir(folder)
        if os.path.splitext(n)[1].lower() in IMAGE_EXTS
    ]
    if not paths:
        sys.exit(f"No image files ({', '.join(sorted(IMAGE_EXTS))}) found in {folder}")
    return sorted(paths, key=natural_key)


# --------------------------------------------------------------------------- #
# Run the engine once per image (this is the "run it multiple times" part)
# --------------------------------------------------------------------------- #

def parse_usage_line(stdout: str) -> dict:
    """Pull the engine's '__USAGE__ {json}' line out of its stdout."""
    import json
    for line in stdout.splitlines():
        if line.startswith("__USAGE__"):
            try:
                return json.loads(line[len("__USAGE__"):].strip())
            except ValueError:
                return {}
    return {}


def run_engine_on_image(python_exe: str, engine_path: str, image: str, out_xlsx: str,
                        provider: str = None, model: str = None) -> dict:
    cmd = [python_exe, engine_path, "--image", image, "--output", out_xlsx, "--no-section-sec"]
    if provider:
        cmd += ["--provider", provider]
    if model:
        cmd += ["--model", model]
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0 or not os.path.isfile(out_xlsx):
        sys.exit(f"Engine failed on {os.path.basename(image)}:\n"
                 f"{result.stdout}\n{result.stderr}")
    return parse_usage_line(result.stdout)


# --------------------------------------------------------------------------- #
# Pure-Python stitching of the per-page .xlsx files
# --------------------------------------------------------------------------- #

def read_page_rows(xlsx_path: str) -> list:
    """Read a per-page Menu sheet into a list of {column: value} dicts."""
    wb = load_workbook(xlsx_path)
    ws = wb["Menu"] if "Menu" in wb.sheetnames else wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = list(next(rows_iter))
    rows = []
    for values in rows_iter:
        rows.append({header[i]: values[i] for i in range(len(header))})
    return rows


def read_page_variants(xlsx_path: str) -> list:
    """Read a per-page Variants sheet into a list of dicts (empty if no sheet)."""
    return _read_named_sheet(xlsx_path, "Variants")


def read_page_modifiers(xlsx_path: str) -> list:
    """Read a per-page Section (modifier) sheet into a list of dicts."""
    return _read_named_sheet(xlsx_path, "Section")


def _read_named_sheet(xlsx_path: str, sheet_name: str) -> list:
    wb = load_workbook(xlsx_path)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = list(next(rows_iter))
    except StopIteration:
        return []
    rows = []
    for values in rows_iter:
        if all(v is None for v in values):
            continue
        rows.append({header[i]: values[i] for i in range(len(header))})
    return rows


def parse_sec(value) -> list:
    """'1,2' -> [1, 2]; handles ints, floats, blanks."""
    if value is None:
        return []
    if isinstance(value, (int, float)):
        return [int(value)]
    out = []
    for token in str(value).split(","):
        token = token.strip()
        if not token:
            continue
        try:
            out.append(int(float(token)))
        except ValueError:
            pass
    return out


def format_sec(ids: list):
    return ",".join(str(i) for i in ids) if ids else None


def _norm(name) -> str:
    return (name or "").strip().lower()


def summarize_page_sections(rows: list, max_items: int = 6) -> list:
    """Reduce a page's rows to [{name, items[]}] for boundary detection."""
    sections = []
    current = None
    for r in rows:
        if r.get("Category ID") is not None:
            current = {"name": r.get("Category"), "items": []}
            sections.append(current)
        if current is None:                       # safety: items before any heading
            current = {"name": None, "items": []}
            sections.append(current)
        name = r.get("Item")
        if name and len(current["items"]) < max_items:
            current["items"].append(name)
    return sections


def stitch(page_paths: list, columns: list, continuation_pages: set):
    """Merge per-page rows into (menu_rows, variant_rows), renumbered.

    `continuation_pages` holds the 0-based indices of pages whose FIRST section
    continues the LAST section of the previous page (so it folds into it instead
    of starting a new Category). That decision is made elsewhere; the merge math
    here is pure Python. Variant rows are re-linked to the renumbered Item ids
    and given fresh globally-unique variant ids.
    """
    merged = []
    variant_rows = []
    section_rows = []
    item_id = 0
    category_id = 0
    sec_offset = 0
    variant_id = 0
    modifier_id = 0
    group_ordinal = 0
    menu_fields = (None, None, None)
    menu_captured = False

    for page_index, path in enumerate(page_paths):
        page_rows = read_page_rows(path)
        if not page_rows:
            continue

        if not menu_captured:
            menu_fields = (page_rows[0].get("Menu Group"),
                           page_rows[0].get("Menu Group 2"),
                           page_rows[0].get("Menu Group 3"))
            menu_captured = True

        local_to_global = {}          # this page's Item ID -> global Item ID
        page_combo_max = 0
        first_section_in_page = True
        for r in page_rows:
            new = {c: None for c in columns}
            new["Item"] = r.get("Item")
            new["Item 2"] = r.get("Item 2")
            new["Item 3"] = r.get("Item 3")
            new["Item Type"] = r.get("Item Type")
            new["Price"] = r.get("Price")

            if r.get("Category ID") is not None:           # a section start on this page
                folds_into_previous = (first_section_in_page
                                       and page_index in continuation_pages
                                       and category_id > 0)
                if not folds_into_previous:
                    category_id += 1
                    new["Category ID"] = category_id
                    new["Category"] = r.get("Category")
                    new["Category 2"] = r.get("Category 2")
                    new["Category 3"] = r.get("Category 3")
                first_section_in_page = False

            # Sec carries combo ids only at this stage (offset so pages don't
            # collide). Section/modifier sec ids are added later by finalize_sec.
            combo_ids = parse_sec(r.get("Sec"))
            if combo_ids:
                page_combo_max = max(page_combo_max, max(combo_ids))
            new["Sec"] = format_sec([i + sec_offset for i in combo_ids])

            item_id += 1
            new["Item ID"] = item_id
            if r.get("Item ID") is not None:
                local_to_global[r.get("Item ID")] = item_id
            if item_id == 1:
                new["Menu ID"] = 1
                new["Menu Group"] = menu_fields[0]
                new["Menu Group 2"] = menu_fields[1]
                new["Menu Group 3"] = menu_fields[2]

            merged.append(new)

        # remap this page's variants onto the global Item ids
        for v in read_page_variants(path):
            gl = local_to_global.get(v.get("Item id"))
            if gl is None:
                continue
            variant_id += 1
            variant_rows.append({
                "Item id":         gl,
                "variant id":      variant_id,
                "variant name":    v.get("variant name"),
                "Language 2 name": v.get("Language 2 name"),
                "price":           v.get("price"),
            })

        # merge this page's section modifiers; renumber each page's group ids to
        # globally-unique ordinals (finalize_sec shifts them past combo ids later).
        page_mods = read_page_modifiers(path)
        local_group_ids = sorted({m.get("Section Id") for m in page_mods
                                  if m.get("Section Id") is not None})
        group_remap = {}
        for lg in local_group_ids:
            group_ordinal += 1
            group_remap[lg] = group_ordinal
        for smod in page_mods:
            modifier_id += 1
            section_rows.append({
                "Section Id":    group_remap.get(smod.get("Section Id")),
                "SectionName":   smod.get("SectionName"),
                "Min":           smod.get("Min"),
                "Max":           smod.get("Max"),
                "Modifier ID":   modifier_id,
                "ModifierName":  smod.get("ModifierName"),
                "ModifierName2": smod.get("ModifierName2"),
                "Price":         smod.get("Price"),
            })

        sec_offset += page_combo_max

    return merged, variant_rows, section_rows


# --------------------------------------------------------------------------- #
# Deciding which page boundaries are "same section, split across pages"
# --------------------------------------------------------------------------- #

def continuations_by_name(page_section_lists: list) -> set:
    """Fallback (no AI): a page continues the previous one if its first section
    name equals the previous page's last section name."""
    cont = set()
    for i in range(1, len(page_section_lists)):
        prev, cur = page_section_lists[i - 1], page_section_lists[i]
        if prev and cur and _norm(cur[0]["name"]) == _norm(prev[-1]["name"]) \
           and _norm(cur[0]["name"]):
            cont.add(i)
    return cont


def _detect_call_claude(prompt: str, model: str):
    import anthropic
    client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))
    response = client.messages.create(
        model=model, max_tokens=500,
        messages=[{"role": "user", "content": prompt}],
    )
    u = getattr(response, "usage", None)
    usage_dict = {
        "model": model,
        "input_tokens": (getattr(u, "input_tokens", 0) or 0) if u else 0,
        "output_tokens": (getattr(u, "output_tokens", 0) or 0) if u else 0,
    }
    text = "".join(b.text for b in (getattr(response, "content", None) or [])
                   if getattr(b, "type", None) == "text")
    return text, usage_dict


def _detect_call_gemini(prompt: str, model: str):
    from google import genai
    from google.genai import types
    client = genai.Client(api_key=(os.environ.get("GEMINI_API_KEY") or os.environ.get("GOOGLE_API_KEY")))
    response = client.models.generate_content(
        model=model, contents=prompt,
        config=types.GenerateContentConfig(
            max_output_tokens=500, temperature=0, response_mime_type="application/json"),
    )
    um = getattr(response, "usage_metadata", None)
    usage_dict = {
        "model": model,
        "input_tokens": (getattr(um, "prompt_token_count", 0) or 0) if um else 0,
        "output_tokens": (((getattr(um, "candidates_token_count", 0) or 0)
                           + (getattr(um, "thoughts_token_count", 0) or 0)) if um else 0),
    }
    text = getattr(response, "text", None) or ""
    return text, usage_dict


def continuations_by_ai(page_section_lists: list, model: str, provider: str = "claude") -> set:
    """Ask the model, in ONE small text-only call, which page boundaries are a
    single section split across the break. Returns 0-based later-page indices.
    Raises on any failure so the caller can fall back."""
    if provider == "gemini":
        if not (os.environ.get("GEMINI_API_KEY") or os.environ.get("GOOGLE_API_KEY")):
            raise RuntimeError("no API key")
    else:
        if not os.environ.get("ANTHROPIC_API_KEY"):
            raise RuntimeError("no API key")

    # Build the list of boundaries between consecutive non-empty pages.
    boundaries = []          # (later_page_index, earlier_last_section, later_first_section)
    for i in range(1, len(page_section_lists)):
        prev, cur = page_section_lists[i - 1], page_section_lists[i]
        if prev and cur:
            boundaries.append((i, prev[-1], cur[0]))
    if not boundaries:
        return set(), {"model": model, "input_tokens": 0, "output_tokens": 0}

    lines = []
    for n, (_, last_sec, first_sec) in enumerate(boundaries, start=1):
        lines.append(
            f"Boundary {n}:\n"
            f"  Earlier page ENDS with section: {last_sec['name']!r}\n"
            f"    last items: {', '.join(last_sec['items']) or '(none)'}\n"
            f"  Later page BEGINS with section: {first_sec['name']!r}\n"
            f"    first items: {', '.join(first_sec['items']) or '(none)'}"
        )
    prompt = (
        "A restaurant menu was photographed one page at a time and each page was "
        "read separately. A single menu SECTION (e.g. \"Lunch Specials\") can be "
        "split across a page break, and the continued part often does NOT reprint "
        "the heading. For each boundary below, decide whether the section that "
        "BEGINS the later page is the SAME section that was still running at the "
        "END of the earlier page (i.e. one section split across the break), rather "
        "than a brand-new section.\n\n"
        + "\n".join(lines)
        + "\n\nReturn ONLY JSON, no prose: {\"continuations\": [list of boundary "
          "numbers that ARE continuations]}."
    )

    if provider == "gemini":
        text, usage_dict = _detect_call_gemini(prompt, model)
    else:
        text, usage_dict = _detect_call_claude(prompt, model)
    data = parse_json_loose(text)
    flagged = {int(x) for x in data.get("continuations", [])}
    later_indices = {boundaries[n - 1][0] for n in flagged if 1 <= n <= len(boundaries)}
    return later_indices, usage_dict


def parse_json_loose(text: str) -> dict:
    import json
    cleaned = text.strip()
    if cleaned.startswith("```"):
        cleaned = cleaned.split("```", 2)[1]
        if cleaned.lstrip().lower().startswith("json"):
            cleaned = cleaned.lstrip()[4:]
    cleaned = cleaned.strip().strip("`").strip()
    return json.loads(cleaned)


# --------------------------------------------------------------------------- #
# Entry point
# --------------------------------------------------------------------------- #

def main() -> None:
    here = os.path.dirname(os.path.abspath(__file__))
    ap = argparse.ArgumentParser(description="Scan a folder of menu pages into one Excel file.")
    ap.add_argument("--folder", required=True, help="Folder of menu page images.")
    ap.add_argument("--output", required=True, help="Path for the combined .xlsx file.")
    ap.add_argument("--engine", default=os.path.join(here, "menu_to_excel_2.py"),
                    help="Path to the per-image extractor (default: menu_to_excel_2.py beside this script).")
    ap.add_argument("--keep-pages", action="store_true", help="Keep the per-page .xlsx files.")
    ap.add_argument("--pages-dir", help="Where to write per-page files (default: a temp folder).")
    ap.add_argument("--python", default=sys.executable, help="Python interpreter for the engine.")
    ap.add_argument("--provider", choices=["claude", "gemini"],
                    help="Model provider for extraction (default: claude).")
    ap.add_argument("--model", help="Override the extraction model (implies provider).")
    ap.add_argument("--detect-model", default=None,
                    help="Model for the cross-page section-continuation check (default: engine's MODEL).")
    ap.add_argument("--no-detect", action="store_true",
                    help="Skip the AI continuation check; merge same-named sections by name only.")
    args = ap.parse_args()

    if args.detect_model is None:
        # default to whatever the engine uses, resolved after the engine loads
        args.detect_model = "__engine__"

    output_path = args.output
    if not output_path.lower().endswith(".xlsx"):
        output_path = os.path.splitext(output_path)[0] + ".xlsx"
        print(f"Note: writing Excel output as {output_path} (not {args.output}).")

    engine = load_engine(args.engine)
    provider, model = engine.resolve_provider_model(args.provider, args.model)
    if args.detect_model == "__engine__":
        args.detect_model = model
    detect_provider = engine.provider_for_model(args.detect_model)
    images = list_images(args.folder)

    pages_dir = args.pages_dir or tempfile.mkdtemp(prefix="menu_pages_")
    os.makedirs(pages_dir, exist_ok=True)

    print(f"Found {len(images)} page(s). Running the extractor ({provider}: {model}) once per page:")
    page_paths = []
    usage_records = []          # one dict per API call, each {model, input_tokens, output_tokens}
    for idx, image in enumerate(images, start=1):
        page_xlsx = os.path.join(pages_dir, f"page_{idx:03d}.xlsx")
        print(f"  [{idx}/{len(images)}] {os.path.basename(image)} -> {os.path.basename(page_xlsx)}")
        usage = run_engine_on_image(args.python, args.engine, image, page_xlsx, provider, model)
        if usage:
            usage.setdefault("model", model)
            usage_records.append(usage)
        page_paths.append(page_xlsx)

    print("Stitching pages together (pure Python)...")
    page_section_lists = [summarize_page_sections(read_page_rows(p)) for p in page_paths]

    if args.no_detect:
        continuation_pages = continuations_by_name(page_section_lists)
        print("  Section continuations across pages: name-matching (AI detection off).")
    else:
        try:
            continuation_pages, detect_usage = continuations_by_ai(
                page_section_lists, args.detect_model, detect_provider)
            if detect_usage:
                usage_records.append(detect_usage)
            print(f"  Section continuations across pages (AI): "
                  f"{sorted(i + 1 for i in continuation_pages) or 'none'}")
        except Exception as exc:
            continuation_pages = continuations_by_name(page_section_lists)
            print(f"  AI continuation check unavailable ({exc}); used name-matching instead.")

    merged_rows, variant_rows, section_rows = stitch(page_paths, engine.COLUMNS, continuation_pages)
    engine.finalize_sec(merged_rows, section_rows)   # assign sec ids + write into Sec
    engine.write_excel(merged_rows, output_path, variant_rows, section_rows)
    print(f"Wrote {len(merged_rows)} items ({len(variant_rows)} size variants, "
          f"{len(section_rows)} section modifiers) from {len(page_paths)} page(s) to {output_path}")

    # ----- money used for the whole run -----
    total_in = sum(u.get("input_tokens", 0) for u in usage_records)
    total_out = sum(u.get("output_tokens", 0) for u in usage_records)
    total_cost = sum(engine.estimate_cost(u.get("model", getattr(engine, "MODEL", "")),
                                          u.get("input_tokens", 0),
                                          u.get("output_tokens", 0))
                     for u in usage_records)
    print("-" * 48)
    if usage_records:
        print(f"API calls:      {len(usage_records)} "
              f"({len(page_paths)} page extraction(s)"
              f"{' + 1 continuation check' if not args.no_detect and len(usage_records) > len(page_paths) else ''})")
        print(f"Tokens used:    {total_in:,} in / {total_out:,} out")
        print(f"Total cost:     ${total_cost:.4f}  (estimate at current per-million rates)")
    else:
        print("Total cost:     unavailable (no usage was reported by the extractor).")

    if not args.keep_pages and not args.pages_dir:
        for p in page_paths:
            try:
                os.remove(p)
            except OSError:
                pass
        try:
            os.rmdir(pages_dir)
        except OSError:
            pass
    else:
        print(f"Per-page files kept in: {pages_dir}")


if __name__ == "__main__":
    main()
