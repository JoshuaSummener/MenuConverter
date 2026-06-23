#!/usr/bin/env python3
"""
to_pos_format.py
================
Convert the scanner's combined workbook (Menu / Variants / Section sheets) into
the normalized POS import workbook, enforcing the import system's validation
rules so the file passes the preview check.

No template file is required: the POS format (all 11 sheets + headers) is built
in-code. You may optionally pass --template to use the vendor's exported file as
the base instead (which also carries its README text and _Lookups formulas).

Full chain:
    folder of photos
        -> menu_folder_to_excel.py   (-> our combined .xlsx)
        -> to_pos_format.py          (-> POS import .xlsx)

Usage
-----
    python to_pos_format.py --input our_menu.xlsx --output pos.xlsx
    python to_pos_format.py --input our_menu.xlsx --output pos.xlsx --template store-menu-TEMPLATE.xlsx
    python to_pos_format.py ... --menu-name "China Taste"

Validation rules enforced (from the import guide):
  Names      Item <=50 (overflow kept in Description), Modifier <=50, Option <=100,
             Category <=100, Group/Menu <=255, Description <=500.
  Numbers    prices >=0 (item price <=999,999.99); Min >=0; Max >=1 and >=Min;
             Markup 0-100.
  Status     ACTIVE.
  Counts     warns if a menu has >50 groups or a group has >200 items.
  Sort       unique within scope (groups within menu; items within group).
"""

import argparse
import os
import sys

from openpyxl import Workbook, load_workbook

MAX_ITEM_NAME = 50
MAX_MODIFIER_NAME = 50
MAX_OPTION_NAME = 100
MAX_CATEGORY_NAME = 100
MAX_GROUP_NAME = 255
MAX_MENU_NAME = 255
MAX_DESCRIPTION = 500
MAX_ITEM_PRICE = 999999.99
MAX_GROUPS_PER_MENU = 50
MAX_ITEMS_PER_GROUP = 200

# --------------------------------------------------------------------------- #
# Built-in POS schema (so no external template file is required)
# --------------------------------------------------------------------------- #

SHEET_ORDER = ["README", "Menus", "Menu Groups", "Menu Items", "Menu Modifier",
               "Schedule", "Items", "Modifiers", "Disabled Options", "Categories", "_Lookups"]

BASE_HEADERS = {
    "Menus": ['Id*', 'Menu Name*', 'Description', 'Sort', 'Markup %', 'is_pos*', 'is_online*', 'is_orderai*', 'Status*', 'Menu Name (ar)', 'Menu Name (en)', 'Menu Name (es)', 'Menu Name (ja)', 'Menu Name (vi)', 'Menu Name (zh-Hans)', 'Menu Name (zh-Hant)', 'Description (ar)', 'Description (en)', 'Description (es)', 'Description (ja)', 'Description (vi)', 'Description (zh-Hans)', 'Description (zh-Hant)', 'Sync Id'],
    "Menu Groups": ['Id*', 'Menu Id*', 'Group Name*', 'Sort', 'Group Name (ar)', 'Group Name (en)', 'Group Name (es)', 'Group Name (ja)', 'Group Name (vi)', 'Group Name (zh-Hans)', 'Group Name (zh-Hant)', 'Sync Id'],
    "Menu Items": ['Menu Id*', 'Group Id*', 'Item Id*', 'Item Name (auto)', 'Sort', 'Price (auto)', 'Menu Price', 'Sync Id'],
    "Menu Modifier": ['Menu Id*', 'Modifier Id*', 'Option Id*', 'Modifier Name (auto)', 'Option Name (auto)', 'Price (auto)', 'Menu Price', 'Sync Id'],
    "Items": ['Id*', 'Name*', 'Description', 'Price*', 'Stock', 'Status*', 'Category Ids', 'Modifier Ids', 'Image URL', 'Name (ar)', 'Name (en)', 'Name (es)', 'Name (ja)', 'Name (vi)', 'Name (zh-Hans)', 'Name (zh-Hant)', 'Description (ar)', 'Description (en)', 'Description (es)', 'Description (ja)', 'Description (vi)', 'Description (zh-Hans)', 'Description (zh-Hant)', 'Sync Id'],
    "Modifiers": ['Modifier Id*', 'Modifier Name*', 'Min Selection', 'Max Selection', 'Option Id*', 'Option Name*', 'Price*', 'Sort', 'Modifier Name (ar)', 'Modifier Name (en)', 'Modifier Name (es)', 'Modifier Name (ja)', 'Modifier Name (vi)', 'Modifier Name (zh-Hans)', 'Modifier Name (zh-Hant)', 'Option Name (ar)', 'Option Name (en)', 'Option Name (es)', 'Option Name (ja)', 'Option Name (vi)', 'Option Name (zh-Hans)', 'Option Name (zh-Hant)', 'Modifier Sync Id', 'Option Sync Id'],
    "Disabled Options": ['Item Id*', 'Modifier Id*', 'Option Id*', 'Item Name (auto)', 'Modifier Name (auto)', 'Option Name (auto)', 'Disabled?', 'Sync Id'],
    "Categories": ['Id*', 'Name*', 'Sort', 'Description', 'Image URL', 'Name (ar)', 'Name (en)', 'Name (es)', 'Name (ja)', 'Name (vi)', 'Name (zh-Hans)', 'Name (zh-Hant)', 'Description (ar)', 'Description (en)', 'Description (es)', 'Description (ja)', 'Description (vi)', 'Description (zh-Hans)', 'Description (zh-Hant)', 'Sync Id'],
}
SCHEDULE_ROWS = [['Store Timezone (read-only):', 'America/New_York'], [],
                 ['Menu Id*', 'Group Id (blank=whole menu)', 'Day*', 'Start*', 'End*']]
LOOKUPS_HEADER = ['Menu picker', 'Menu Group picker', 'Item picker', 'Modifier picker', 'Option picker', 'Category picker']


def build_base_workbook():
    """Create an empty POS workbook with the correct sheets and headers, so the
    converter needs no external template file."""
    wb = Workbook()
    wb.remove(wb.active)
    for name in SHEET_ORDER:
        ws = wb.create_sheet(name)
        if name == "README":
            ws["A1"] = "Menu Import workbook (generated). Fill the data tabs and import."
        elif name == "Schedule":
            for r, row in enumerate(SCHEDULE_ROWS, start=1):
                for c, val in enumerate(row, start=1):
                    ws.cell(row=r, column=c, value=val)
        elif name == "_Lookups":
            for c, val in enumerate(LOOKUPS_HEADER, start=1):
                ws.cell(row=1, column=c, value=val)
        else:
            for c, val in enumerate(BASE_HEADERS[name], start=1):
                ws.cell(row=1, column=c, value=val)
    return wb


# --------------------------------------------------------------------------- #
# Reading OUR combined workbook
# --------------------------------------------------------------------------- #

def header_map(ws):
    return {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}


def read_our_workbook(path):
    wb = load_workbook(path, data_only=True)
    menu_ws = wb["Menu"]
    h = header_map(menu_ws)

    menu_name = None
    items = []
    cur_category = None
    for r in range(2, menu_ws.max_row + 1):
        def g(col):
            idx = h.get(col)
            return menu_ws.cell(row=r, column=idx).value if idx else None
        if g("Menu Group"):
            menu_name = menu_name or g("Menu Group")
        if g("Category"):
            cur_category = g("Category")
        item_id = g("Item ID")
        if item_id is None:
            continue
        items.append({
            "id": int(item_id),
            "english": g("Item") or "",
            "lang2": g("Item 2") or "",
            "price": float(g("Price") or 0),
            "category": cur_category or "Menu",
        })
    menu_name = menu_name or "Menu"

    variants = {}
    if "Variants" in wb.sheetnames:
        vws = wb["Variants"]; vh = header_map(vws)
        for r in range(2, vws.max_row + 1):
            iid = vws.cell(row=r, column=vh["Item id"]).value
            if iid is None:
                continue
            variants.setdefault(int(iid), []).append((
                vws.cell(row=r, column=vh["variant name"]).value,
                vws.cell(row=r, column=vh["Language 2 name"]).value,
                float(vws.cell(row=r, column=vh["price"]).value or 0),
            ))

    section_groups = []
    if "Section" in wb.sheetnames:
        sws = wb["Section"]; sh = header_map(sws)
        cur_name = None
        groups = {}
        order = []
        for r in range(2, sws.max_row + 1):
            name = sws.cell(row=r, column=sh["SectionName"]).value
            if name:
                cur_name = name
            modname = sws.cell(row=r, column=sh["ModifierName"]).value
            if modname is None:
                continue
            gmin = int(sws.cell(row=r, column=sh["Min"]).value or 0)
            gmax = int(sws.cell(row=r, column=sh["Max"]).value or 1)
            key = (cur_name, gmin, gmax)
            if key not in groups:
                groups[key] = {"section": cur_name, "min": gmin, "max": gmax, "options": []}
                order.append(key)
            groups[key]["options"].append((
                modname,
                sws.cell(row=r, column=sh["ModifierName2"]).value,
                float(sws.cell(row=r, column=sh["Price"]).value or 0),
            ))
        section_groups = [groups[k] for k in order]

    return menu_name, items, variants, section_groups


# --------------------------------------------------------------------------- #
# Validation helpers
# --------------------------------------------------------------------------- #

class Adjust:
    def __init__(self):
        self.truncated = 0
        self.price_clamped = 0
        self.sel_clamped = 0
        self.warnings = []

    def clip(self, text, maxlen):
        if text is None:
            return text
        t = str(text)
        if len(t) > maxlen:
            self.truncated += 1
            return t[:maxlen].rstrip()
        return t

    def price(self, p):
        try:
            p = float(p)
        except (TypeError, ValueError):
            p = 0.0
        out = min(max(p, 0.0), MAX_ITEM_PRICE)
        if abs(out - p) > 1e-9:
            self.price_clamped += 1
        return round(out, 2)

    def selection(self, gmin, gmax):
        nmin = max(int(gmin or 0), 0)
        nmax = max(int(gmax or 1), 1, nmin)
        if nmin != (gmin or 0) or nmax != (gmax or 1):
            self.sel_clamped += 1
        return nmin, nmax


def bilingual(english, lang2):
    english = (english or "").strip()
    lang2 = (lang2 or "").strip()
    return f"{lang2} {english}".strip() if lang2 else english


def ref(id_, name):
    return f"{id_} - {name}"


def clear_data(ws):
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)


# --------------------------------------------------------------------------- #
# Build the POS workbook
# --------------------------------------------------------------------------- #

def convert(our_path, template_path, out_path, menu_name_override=None):
    adj = Adjust()
    menu_name, items, variants, section_groups = read_our_workbook(our_path)
    if menu_name_override:
        menu_name = menu_name_override
    menu_name = adj.clip(menu_name or "Menu", MAX_MENU_NAME)

    # Categories (first-appearance order)
    cat_order = []
    for it in items:
        if it["category"] and it["category"] not in cat_order:
            cat_order.append(it["category"])
    cat_id = {name: i + 1 for i, name in enumerate(cat_order)}

    # Modifiers: section groups first, then per-item size groups
    modifiers = []
    option_seq = 0
    mod_seq = 0
    cat_modifier_ids = {}

    for grp in section_groups:
        mod_seq += 1
        gmin, gmax = adj.selection(grp["min"], grp["max"])
        label = adj.clip((grp["section"] or "Options") + (" Choice" if gmin >= 1 else " Add-ons"), MAX_MODIFIER_NAME)
        opts = []
        for (oname, ol2, price) in grp["options"]:
            option_seq += 1
            opts.append((option_seq, adj.clip(bilingual(oname, ol2) or f"Option {option_seq}", MAX_OPTION_NAME), adj.price(price)))
        modifiers.append({"mid": mod_seq, "name": label, "min": gmin, "max": gmax, "options": opts})
        cat_modifier_ids.setdefault(grp["section"], []).append(mod_seq)

    item_size_modifier = {}
    item_base_price = {}
    for it in items:
        vs = variants.get(it["id"])
        if not vs:
            continue
        base = min(p for (_, _, p) in vs)
        item_base_price[it["id"]] = base
        mod_seq += 1
        opts = []
        for (vname, vl2, price) in vs:
            option_seq += 1
            opts.append((option_seq, adj.clip(bilingual(vname, vl2) or f"Option {option_seq}", MAX_OPTION_NAME), adj.price(price - base)))
        modifiers.append({"mid": mod_seq, "name": "Size", "min": 1, "max": 1, "options": opts})
        item_size_modifier[it["id"]] = mod_seq

    # per-item name/price/modifier ids, with Item Name <=50 (overflow -> Description)
    for it in items:
        full = bilingual(it["english"], it["lang2"])
        it["name"] = adj.clip(full or f"Item {it['id']}", MAX_ITEM_NAME)
        it["description"] = adj.clip(full, MAX_DESCRIPTION) if len(full) > MAX_ITEM_NAME else ""
        it["base_price"] = adj.price(item_base_price.get(it["id"], it["price"]))
        ids = list(cat_modifier_ids.get(it["category"], []))
        if it["id"] in item_size_modifier:
            ids.append(item_size_modifier[it["id"]])
        it["modifier_ids"] = ids

    # count limits
    if len(cat_order) > MAX_GROUPS_PER_MENU:
        adj.warnings.append(f"{len(cat_order)} groups exceeds the {MAX_GROUPS_PER_MENU}-per-menu limit.")
    per_group = {}
    for it in items:
        per_group[it["category"]] = per_group.get(it["category"], 0) + 1
    for cat, n in per_group.items():
        if n > MAX_ITEMS_PER_GROUP:
            adj.warnings.append(f"group '{cat}' has {n} items (>{MAX_ITEMS_PER_GROUP}).")

    # ----- load template, or build the structure in-code if none given -----
    if template_path and os.path.exists(template_path):
        wb = load_workbook(template_path)
    else:
        wb = build_base_workbook()
    for s in ["Menus", "Menu Groups", "Menu Items", "Menu Modifier",
              "Items", "Modifiers", "Categories", "Disabled Options"]:
        if s in wb.sheetnames:
            clear_data(wb[s])
    # clear the template's example Schedule rows (they referenced the old menu)
    if "Schedule" in wb.sheetnames:
        sch = wb["Schedule"]
        for r in range(1, sch.max_row + 1):
            if sch.cell(row=r, column=1).value == "Menu Id*" and sch.max_row > r:
                sch.delete_rows(r + 1, sch.max_row - r)
                break

    def put(ws, row_idx, hi, mapping):
        for col_name, value in mapping.items():
            if col_name in hi:
                ws.cell(row=row_idx, column=hi[col_name], value=value)

    ws = wb["Categories"]; hi = header_map(ws)
    for i, name in enumerate(cat_order):
        put(ws, i + 2, hi, {"Id*": cat_id[name], "Name*": adj.clip(name, MAX_CATEGORY_NAME), "Sort": i + 1})

    ws = wb["Items"]; hi = header_map(ws)
    for i, it in enumerate(items):
        put(ws, i + 2, hi, {
            "Id*": it["id"], "Name*": it["name"], "Description": it["description"],
            "Price*": it["base_price"], "Status*": "ACTIVE",
            "Category Ids": str(cat_id[it["category"]]) if it["category"] in cat_id else "",
            "Modifier Ids": ", ".join(str(m) for m in it["modifier_ids"]),
        })

    ws = wb["Modifiers"]; hi = header_map(ws)
    r = 2
    for m in modifiers:
        for sort_i, (oid, oname, price) in enumerate(m["options"]):
            put(ws, r, hi, {
                "Modifier Id*": m["mid"], "Modifier Name*": m["name"],
                "Min Selection": m["min"], "Max Selection": m["max"],
                "Option Id*": oid, "Option Name*": oname, "Price*": price, "Sort": sort_i,
            })
            r += 1

    ws = wb["Menus"]; hi = header_map(ws)
    put(ws, 2, hi, {"Id*": 1, "Menu Name*": menu_name, "Sort": 0, "Markup %": 0,
                    "is_pos*": "TRUE", "is_online*": "TRUE", "is_orderai*": "TRUE", "Status*": "ACTIVE"})
    menu_ref = ref(1, menu_name)

    ws = wb["Menu Groups"]; hi = header_map(ws)
    group_id = {}
    for i, name in enumerate(cat_order):
        gid = i + 1
        group_id[name] = gid
        put(ws, i + 2, hi, {"Id*": gid, "Menu Id*": menu_ref,
                            "Group Name*": adj.clip(name, MAX_GROUP_NAME), "Sort": i + 1})

    ws = wb["Menu Items"]; hi = header_map(ws)
    sort_per_group = {}
    for i, it in enumerate(items):
        cat = it["category"]
        s = sort_per_group.get(cat, 0); sort_per_group[cat] = s + 1
        put(ws, i + 2, hi, {
            "Menu Id*": menu_ref,
            "Group Id*": ref(group_id.get(cat, ""), adj.clip(cat, MAX_GROUP_NAME)) if cat else "",
            "Item Id*": ref(it["id"], it["name"]),
            "Sort": s,
        })

    ws = wb["Menu Modifier"]; hi = header_map(ws)
    r = 2
    for m in modifiers:
        for (oid, oname, price) in m["options"]:
            put(ws, r, hi, {"Menu Id*": menu_ref,
                            "Modifier Id*": ref(m["mid"], m["name"]),
                            "Option Id*": ref(oid, oname)})
            r += 1

    wb.save(out_path)

    print(f"Categories: {len(cat_order)} | Items: {len(items)} | "
          f"Modifiers: {len(modifiers)} | Options: {option_seq}")
    print(f"Validation adjustments: {adj.truncated} name(s) truncated, "
          f"{adj.price_clamped} price(s) clamped, {adj.sel_clamped} min/max clamped.")
    for w in adj.warnings:
        print(f"  WARNING: {w}")
    print(f"Saved POS workbook -> {out_path}")


def main():
    ap = argparse.ArgumentParser(description="Convert the scanner workbook into the POS import format.")
    ap.add_argument("--input", required=True, help="Our combined .xlsx (Menu/Variants/Section sheets).")
    ap.add_argument("--template", help="Optional POS template .xlsx to use as the base. "
                                       "If omitted, the format is built in-code (no file needed).")
    ap.add_argument("--output", required=True, help="Path for the POS import .xlsx to write.")
    ap.add_argument("--menu-name", help="Override the menu name (default: taken from the workbook).")
    args = ap.parse_args()
    try:
        convert(args.input, args.template, args.output, args.menu_name)
    except KeyError as exc:
        sys.exit(f"Missing expected sheet/column: {exc}")


if __name__ == "__main__":
    main()
